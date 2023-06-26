# -*- python -*-
"""This library is used to maintain header content on a spreadsheet."""

##-------------------##
##---]  GLOBALS  [---##
##-------------------##

##-------------------##
##---]  IMPORTS  [---##
##-------------------##
import pprint

from openpyxl.styles         import Alignment
from openpyxl.styles.borders import Border, Side

from stats.spreadsheet import styles       as sheet_styles
from openpyxl.worksheet.worksheet \
                       import Worksheet

from vlan.main          import utils       as main_utils
from stats.spreadsheet import cells        as onf_cells


class Utils:
    '''.'''

    wb  = None

    ## -----------------------------------------------------------------------
    ## -----------------------------------------------------------------------
    def __init__(self, wb, tab:str):
        """Constructor.

        :param wb: A spreadsheet workbook to perform actions on.
        :type  wb: openpyxl.Workbook()
        """

        self.wb  = wb
        self.tab = tab

    ## -----------------------------------------------------------------------
    ## -----------------------------------------------------------------------
    def get_summaries(self):
        return ['TOTAL', 'SUCCESS', 'FAILURE', 'UNSTABLE']
        
    ## -----------------------------------------------------------------------
    ## -----------------------------------------------------------------------
    def get_min_max_es(self):
        return ['max', 'avg', 'jobs', 'url', '__SEP__']
        
    ## -----------------------------------------------------------------------
    ## -----------------------------------------------------------------------
    def generate_tab(self):
        tab  = self.tab
        work = self.wb

        self.create_tab(tab)

        summaries  = self.get_summaries()
        min_max_es = self.get_min_max_es()
        min_max_len = len(min_max_es)

        col = 1
        for summary in summaries:
            self.col_hdr(summary, col=col, row=1, merge=min_max_len)
            for min_max in min_max_es:
                if min_max == '__SEP__':
                    self.col_hdr('', col=col, row=2, width=1)
                else:
                    self.col_hdr(min_max, col=col, row=2, width=5)
                col = col + 1

        headers=\
            [
                { 'label':'View',  'width':25, },
                { 'label':'Suite', 'width':70, },
                { 'label':'Notes', },
            ]

        for idx in range(1,50):
            headers += [ {'label':''} ] # extend header offscreen

        for header in headers:
            self.col_hdr('', col=col, row=1)
            width = header['width'] if 'width' in header else None
            self.col_hdr(header['label'], col=col, row=2, width=width)
            col = col + 1

    ## -----------------------------------------------------------------------
    ## -----------------------------------------------------------------------
    def create_tab(self, tab:str) -> Worksheet:
        '''.'''

        sheet = self.wb.create_sheet(title=tab)
        self.wb.active = sheet
        return sheet

    ## -----------------------------------------------------------------------
    ## -----------------------------------------------------------------------
    def get_tab(self, tab:str) -> Worksheet:
        '''.'''

        # sheet = self.wb.get_sheet_by_name(tab)
        sheet = self.wb[tab]
        self.wb.active = sheet
        return sheet

    ## -----------------------------------------------------------------------
    ## -----------------------------------------------------------------------
    def col_hdr(self, word,     \
                align=None,     \
                col:int=None,   \
                merge:int=None, \
                row:int=None,   \
                tab:str=None ,  \
                width:int=None, \
                ):
        '''
        :param word: Text to place in a spreadsheet cell.
        :type  word: str

        :param merge: Number of cells to merge for this header cell.
        :type  merge: int

        :param width: Cell width for displaying text.
        :type  width: int
        '''

        sheet = self.get_tab(tab)   \
            if tab is not None else \
            self.wb.active

        if col is None:
            raise ValueError("col= is invalid")
        
        if row is None:
            row = 1

        cell = sheet.cell(row=row, column=col)
        # cell.border = Border(outline=True)

        cell.border = sheet_styles.get_border('thin')
        cell.font   = sheet_styles.get_fonts('yellow')
        cell.fill   = sheet_styles.get_fills('maroon')

        if align is None:
            cell.alignment = sheet_styles.get_aligns('center')

        if width is not None:
            letter = onf_cells.Posn().get_letter(col)
            sheet.column_dimensions[letter].width = width

        if merge is not None:
            col_range = onf_cells.Posn(sheet)\
                .col_range(row, col, merge-1)
            sheet.merge_cells(col_range) # 'C1:F1'

        cell.value = word

    ## -----------------------------------------------------------------------
    ## -----------------------------------------------------------------------  
    def get_headers(self, raw:list=None):
        """Given basic spreadsheet header data populate the structure with defaults.

        :param raw: header data to format (label=, width=)
        :type  raw: dict

        :return: Header values suitble for spreadsheet rendering.
        :rtype : dict 
           align   left, center, right
           col     Index of column to render
           label   Column label
           skip    Exclude line item from rendering.
           width   Explicit column width (width=None will render to largest data value).
        """

        if raw is None:
            raw = self.get_header_raw()

        def_none = ['align']
        def_bool = ['skip']
        def_list = []
        def_list += def_none
        def_list += def_bool

        # ------------
        # Add defaults
        # ------------ 
        headers = [] # accumulate non-skip records
        col = 1
        for header in raw:

            if 'skip' in header and header['skip']:
                continue

            letter = onf_cells.Posn().get_letter(col)            
            header['col']        = col 
            header['col-letter'] = letter
            
            # Apply defaults
            for key in def_list:
                if key in header.keys():
                    continue
                elif key in def_bool:
                    header[key] = False
                elif key in def_none:
                    header[key] = None
                    
            headers += [header]
            col = col + 1
                
        return headers

    ## -----------------------------------------------------------------------
    ## -----------------------------------------------------------------------  
    def set_row(self, row:int, data:dict):
        '''Populate a spreadsheet row with data.'''

        tab = self.tab
        sheet = self.wb[tab]

        if 2 > row:
            iam = main_utils.iam()
            raise ValueError("%s: row=%s is invalid" % (iam, row))

        col = 1        
        summaries  = self.get_summaries()
        min_max_es = self.get_min_max_es()

        for summary in summaries:
            for min_max in min_max_es:
                cell = sheet.cell(row=row, column=col)
                cell.alignment = sheet_styles.get_aligns('center')
                value = None
                if min_max == '__SEP__':
                    letter = onf_cells.Posn().get_letter(col)
                    sheet.column_dimensions[letter].width = .2
                    cell.fill  = sheet_styles.get_fills('maroon')
                    value = ''
                else:
                    value = data[summary][min_max]

                cell.value = value
                col = col + 1

        for field in ['view', 'suites']:
            if field != 'separator':
                cell = sheet.cell(row=row, column=col)
                cell.alignment = sheet_styles.get_aligns('left')
                cell.value = data[field]
            col = col + 1
            
# [SEE ALSO]
# -----------------------------------------------------------------------
# .. seealso: https://openpyxl.readthedocs.io/en/latest/api/openpyxl.cell.cell.html
# -----------------------------------------------------------------------
       
# [EOF]
