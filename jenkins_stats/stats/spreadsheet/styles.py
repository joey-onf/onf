# -*- python -*-
## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
""" . """

##-------------------##
##---]  IMPORTS  [---##
##-------------------##
import pprint

from stats.spreadsheet  import colors         as onf_colors

from openpyxl.styles    import Alignment
from openpyxl.styles.borders\
                        import Border, Side
from openpyxl.styles    import Color, PatternFill, Font, Border
from openpyxl.styles    import colors

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_aligns(val=None):
    orients = ['left', 'center', 'right'] # 'justify'
    if val is not None and val not in orients:
        raise ValueError("get_aligns: detected invalid alignment %s" % val)

    align = { orient : Alignment(horizontal=orient,vertical='center') for orient in orients }
    ans = align[val] if val is not None else align
    return ans

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_border(name=None):
    '''Return a border definition used to format a spreadsheet cell.

    .. seealso: https://stackoverflow.com/questions/24917201/applying-borders-to-a-cell-in-openpyxl
    .. seealso: https://openpyxl.readthedocs.io/en/latest/_modules/openpyxl/styles/borders.html
    '''

    opts    = {}
    borders = ['double', 'thin', 'thick']

    for border in borders:
        opts[border] = Border\
             (
                 left=Side(style=border), 
                 right=Side(style=border), 
                 top=Side(style=border), 
                 bottom=Side(style=border),
             )

    opts['ugly'] = Border\
        (
            left    = Side(style='thick'), 
            right   = Side(style='thick'), 
            top     = Side(style='thick'), 
            bottom  = Side(style='thick'),
            # color   = 'D4D977',             # colors.get_puke_green()
            outline = True,
        )

    default = 'thin'
    if name is None:
        name = default

    ans = opts[name] if name in opts else opts[ugly]
    return ans
    
## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_fills(color=None):

    # ans1 = PatternFill(fill_type='solid', start_color='7C3A5B')
    hexcode = onf_colors.Hue().get_color_hex(color=color)
    ans = PatternFill(fill_type='solid', start_color=hexcode)
    return ans


    redFill = PatternFill(fill_type='solid', start_color='00FF0000')
    darkOrchidFill = PatternFill(fill_type='solid', start_color='9932CC')
#    orchidB5 = PatternFill(fill_type='solid', start_color='B570EE')
#     magenta  = PatternFill(fill_type='solid', start_color='7C3A5B')
    # purpleFill = PatternFill(fill_type='solid', start_color='C694F8')
    # purpleFill = PatternFill(fill_type='solid', start_color='AE7CE0') # light
    purpleFill = PatternFill(fill_type='solid', start_color='5F2D91')
    # purpleFill = PatternFill(fill_type='solid', start_color='9664C8')
    # https://hexcolorcodes.org/hex-code/9664c8    

    # 00993366

    colors=\
        {
            'dark_orchid' : '9932CC',
            'medium_orchid' : 'BA55D3',
            'orchidB5'  : 'B570EE',
            #
            'maroon'  : '7C3A5B',
            #
            'orchid_dark'   : '9932CC',
            'orchid_medium' : 'BA55D3',
            'orchid_Br'     : 'B570EE',
            #
            'purple'  : '5F2D91',
            'purple1' : 'AD2089',
            #
            'red'     : '00FF0000',
            'orange'  : '00FF9000',
            # 'orange'  : 'FFBB00',
            # 'orange' : '00FFBF00',
            # 'dark-orange' : '00D56F00',
            # 'pink'   : '00FFB9C3',
            'pink'   : '00ffe9ec',
            # 'pink'   : '00ff7386',
            'grey'   : 'ecececec',
            # 
            'yellow'  : 'FAFA00',
            'yellow1' : 'FAFA00',
            'yellow2' : 'FFFF48',
        }

    fill = {}
    for label,hexcode in colors.items():
        fill[label] = PatternFill(fill_type='solid', start_color=hexcode),
        

#    fill=\
#        {
#            'dark_orchid' : PatternFill(fill_type='solid', start_color='9932CC'),
#            'medium_orchid' : PatternFill(fill_type='solid', start_color='BA55D3'),
#            'orchidB5'  : PatternFill(fill_type='solid', start_color='B570EE'),
##            #
 #           'maroon'  : PatternFill(fill_type='solid', start_color='7C3A5B'),
 #           #
 #           'orchid_dark'   : PatternFill(fill_type='solid', start_color='9932CC'),
 #           'orchid_medium' : PatternFill(fill_type='solid', start_color='BA55D3'),
 #           'orchid_Br'     : PatternFill(fill_type='solid', start_color='B570EE'),
 #           #
 #           'purple'  : PatternFill(fill_type='solid', start_color='5F2D91'),
 #           'purple1' : PatternFill(fill_type='solid', start_color='AD2089'),
 #           #
 #           'red'     : PatternFill(fill_type='solid', start_color='00FF0000'),
 #           'orange'  : PatternFill(fill_type='solid', start_color='00FF9000'),
 #           # 'orange' : PatternFill(fill_type='solid', start_color='00FFBF00'),
 #           # 'dark-orange' : PatternFill(fill_type='solid', start_color='00D56F00'),
 #           # 'pink'   : PatternFill(fill_type='solid', start_color='00FFB9C3'),
 #           'pink'   : PatternFill(fill_type='solid', start_color='00ffe9ec'),
 #           # 'pink'   : PatternFill(fill_type='solid', start_color='00ff7386'),
 #           'grey'   : PatternFill(fill_type='solid', start_color='ecececec'),           
 #       }

    # gradient = PatternFill(fill_type=None,
    # start_color='FFFFFFFF',
    # end_color='FF000000')    

    ans = fill[val] if val is not None else fill
    return ans

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_fonts(val=None) -> dict:

    color=\
        {
            'orange'  : 'FFBB00',
            'yellow'  : 'FAFA00',
            'yellow1' : 'FAFA00',
            'yellow2' : 'FFFF48',
        }

    fonts=\
        {
            color : Font(size=10,
                         color=hex_code, \
                         bold=True, \
                         )
            for color,hex_code in color.items()
        }

    ans = fonts[val] if val is not None else fonts
    return ans

# [EOF]
