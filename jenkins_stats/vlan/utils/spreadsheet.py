# -*- python -*-
## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
""" . """

##-------------------##
##---]  IMPORTS  [---##
##-------------------##
import pprint

from pathlib            import PurePath, Path
import time

from openpyxl           import Workbook

from openpyxl.cell      import Cell
from openpyxl.utils     import FORMULAE
from openpyxl.worksheet.hyperlink import Hyperlink, HyperlinkList

# from openpyxl.worksheet.cell_range import CellRange

from openpyxl.styles    import Alignment
from openpyxl.styles    import Color, PatternFill, Font, Border
from openpyxl.styles    import colors

from vlan.main          import argparse        as main_getopt

from vlan.jenkins       import jobs_by

from vlan.spreadsheet   import hdr, col, row

from vlan.utils.consts  import \
    get_types_result,          \
    get_types_state,           \
    get_types_stat

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_letter(col, row=None):
    """."""

    if not isinstance(col, int):
        col = get_column_idx(col)
        if not isinstance(col, int):
            raise ValueError("col=[%s] is invalid" % col)
        
    letter = chr(ord('A') + col - 1)
    if row:
        letter = "%s%d" % (letter, row)
    return letter

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_view_tabs():
    """Return a list of shortented view names."""

    ans  = []
    argv = main_getopt.get_argv()

    if len(argv['view_name']) == 0:
        # Expensive...
        fields = jobs_by.IndexUtils().view_to_jobs()
        argv['view_name'] += list(fields.keys())
        
    views = sorted(argv['view_name'])
    for idx, view in enumerate(views):
        vx = 'v%s' % (idx)
        ans += [ (idx, vx, view) ]

    return ans

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_aligns(val=None):
    orients = ['left', 'center', 'right'] # 'justify'
    if val is not None and val not in orients:
        raise ValueError("get_aligns: detected invalid alignment %s" % val)

    align = { orient : Alignment(horizontal=orient,vertical='center') for orient in orients }
    ans = align[val] if val is not None else align
    return ans

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_fills(val=None):
    redFill = PatternFill(fill_type='solid', start_color='00FF0000')
    darkOrchidFill = PatternFill(fill_type='solid', start_color='9932CC')
    # purpleFill = PatternFill(fill_type='solid', start_color='C694F8')
    # purpleFill = PatternFill(fill_type='solid', start_color='AE7CE0') # light
    purpleFill = PatternFill(fill_type='solid', start_color='5F2D91')
    # purpleFill = PatternFill(fill_type='solid', start_color='9664C8')
    # https://hexcolorcodes.org/hex-code/9664c8    

    fill=\
        {
            'dark_orchid' : PatternFill(fill_type='solid', start_color='9932CC'),
            'purple' : PatternFill(fill_type='solid', start_color='5F2D91'),
            'red'    : PatternFill(fill_type='solid', start_color='00FF0000'),
            'orange' : PatternFill(fill_type='solid', start_color='00FF9000'),
            # 'orange' : PatternFill(fill_type='solid', start_color='00FFBF00'),
            # 'dark-orange' : PatternFill(fill_type='solid', start_color='00D56F00'),
            # 'pink'   : PatternFill(fill_type='solid', start_color='00FFB9C3'),
            'pink'   : PatternFill(fill_type='solid', start_color='00ffe9ec'),
            # 'pink'   : PatternFill(fill_type='solid', start_color='00ff7386'),
            'grey'   : PatternFill(fill_type='solid', start_color='ecececec'),           
        }

    
    # gradient = PatternFill(fill_type=None,
    # start_color='FFFFFFFF',
    # end_color='FF000000')    

    ans = fill[val] if val is not None else fill
    return ans

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_fonts(val=None) -> dict:

    color=\
        {
            'orange'  : 'FFBB00',
            'yellow'  : 'FAFA00',
            'yellow1' : 'FAFA00',
            'yellow2' : 'FFFF48',
        }

    fonts=\
        {
            color : Font(size=10, underline='single', color=hex_code, bold=True, italic=True)
            for color,hex_code in color.items()
        }

    ans = fonts[val] if val is not None else fonts
    return ans

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
## changelog
    
#    font_color = styles.Color('00FF0000')
#    font = styles.Font(bold=True, color=font_color)
#    side = styles.Side(style=styles.borders.BORDER_THIN)
#    border = styles.Border(top=side, right=side, bottom=side, left=side)
#    alignment = styles.Alignment(horizontal='center', vertical='top')
#    fill_color = styles.Color(rgb='006666FF', tint=0.3)
#    fill = styles.PatternFill(patternType='solid', fgColor=fill_color)

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def col_hdr(sheet, col, word, align=None, row=None, width=None):

    # sheet = Workbook().active
    
    if row is None:
        row = 1

    cell = sheet.cell(row=row, column=col)

    cell.font = get_fonts('yellow')
    cell.fill = get_fills('purple')

    if align is None:
        align = 'center'
    cell.alignment = Alignment(horizontal=align)

    if width is not None:
        letter = chr(ord('A') + col - 1)
        sheet.column_dimensions[letter].width = width
        
    cell.value = word

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def do_sheet_header(ws, headers, col_max=None):
    """Given a data structure render a spreadsheet header.

    :param ws: Worksheet to generate a header for.
    :type  ws:

    :param header: A structure containing header data to render.
    :type  header:
    """

    if col_max is None:
        col_max = 100

    col = 1 # for visiblity in col_fill floop
    for header in headers:
        if not 'col' in header:
            pprint.pprint(header)
            
        col   = header['col']
        label = header['label']
        width = header['width']
        align = header['align']        
        col_hdr(ws, col, label, width=width, align=align)

    for col_fill in range(col+1, col_max):
        col_hdr(ws, col_fill, '')

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------  
def get_raw_header():
    return hdr.HdrUtils().get_headers(raw=None)


## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_header_rec(val=None):

    ans = None
    # headers = get_raw_header()
    headers = hdr.HdrUtils().get_headers(raw=None)

    if val is None:
        ans = headers
    else:
        val_uc = val.upper()
        for header in headers:
            if header['label'] == val_uc:
                ans = header
                break

    if not ans:
        pprint.pprint({
            'iam' : main_utils.iam(),
            'ans' : ans,
        })
        raise ValueError("Invalid header key detected [%s]" % val)

    return ans

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def get_column_idx(val=None):

    hdr = get_header_rec(val)
    ans = hdr['col']
    
    return ans

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def do_header(sheet):
    
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20

    headers = hdr.HdrUtils().get_headers(raw=None)
    # headers = get_raw_header()
    do_sheet_header(sheet, headers)

    # c.fill = PatternFill('gray0625')                  # DOTTED
    # c.fill = PatternFill('solid', fgColor = 'F2F2F2') # SEPARATOR
    # ws['E7'].fill = GradientFill('linear', stop = ('85E4F7','4617F1'))
    # https://pythoninoffice.com/python-openpyxl-excel-formatting-cells/
    return

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def do_stats(ws, row=None):

    return

    if row is None:
        row = 2
    
    for state in get_types_state(): # ABORT, FAIL
        hdr  = get_header_rec(state)
        col = hdr['col']

        for idx in range(0,4):
            cell = ws.cell(row=row, column=col+idx)
            if idx == 0:
                cell.value = state

    return

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def do_header_vertical(sheet, max_row=None):
    """Apply cell based formatting, workaround lack of column-wide apply"""

    if max_row is None:
        max_row = sheet.max_row

    align_def = get_aligns()
 
    headers = hdr.HdrUtils().get_headers(raw=None)
    # headers = get_raw_header()
    for rec in headers:
        label = rec['label']
        col   = get_column_idx(label)

        align = None
        if 'align' not in rec:
            align = 'center'
        elif rec['align'] is None:
            align = 'center'
        else:
            align = rec['align']

        for row_fill in range(3, 1+max_row):
            cell = sheet.cell(row=row_fill, column=col)
            cell.alignment = align_def[align]
    
    return

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def do_formulas(sheet, row):
    """ . """

    total_row = row

    center = Alignment(horizontal='center',vertical='center')

    col = get_column_idx('TOTAL')
    cell = sheet.cell(row=row, column=col)

    total_6   = get_letter('SFUA', 6)
    total_max = get_letter('SFUA', sheet.max_row)
    cell.value = '= COUNTA(%s:%s)' % (total_6, total_max)
    cell = sheet.cell(row=row+1, column=col)

    cell.value = 1
    cell.style = 'Percent'

    states = get_types_result()
    for idx,state in enumerate(states):
        col = get_column_idx(state)
        for row in range(2,4):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = get_aligns('center')

            if row == 2:
                cell.value = '= COUNTIFS(%s:%s, "%s")' % (total_6, total_max, state)
            else:
                numerator   = get_letter(state,   row=total_row)
                denominator = get_letter('TOTAL', row=total_row)

                cell.value = '= %s/%s' % (numerator, denominator)
                cell.style = 'Percent'

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def grey_grid(data):
    """ . """

    color = '#ecececec'
    grey_background = PatternFill(fgColor="ecececec")
    diff_style = DifferentialStyle(fill=grey_background)
    rule = Rule(type="expression", dxf=diff_style)
    rule.formula = ["$H1<3"]
    sheet.conditional_formatting.add("A1:O100", rule)
    
## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def summary_page(wb):

    argv = main_getopt.get_argv()

    ws = wb.get_sheet_by_name('summary')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 70

    raw = [
        { 'label':'view', 'width':8, },
        { 'label':'name', 'width':40, 'align':'left'},
    ]

    headers = hdr.HdrUtils().get_headers(raw)
    do_sheet_header(ws, headers)
   
    row=2
    for rec in get_view_tabs():
        idx, vx, view = rec

        # short view name
        cell = ws.cell(row=idx+row, column=1)
        cell.value = vx
        cell.alignment = get_aligns('center')

        # link-to-view-tab
        cell = ws.cell(row=idx+row, column=2)
        cell.value = view
        cell.hyperlink = '#%s!A1' % vx

## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def apply_per_sheet(wb):
    """Iterate over workbook sheets and apply formatting, totals, etc.

    :param wb: Workbook providing sheets to iterate over.
    :type  wb: openpyxl.workbook.Workbook
    """

    # ------------------------------
    # Apply to each sheet with stats
    # ------------------------------
    hdr_row = 2
    for sheet in wb.worksheets:
        if sheet.title == 'summary':
            continue

        do_formulas(sheet, hdr_row)
        do_header_vertical(sheet)
        do_stats(sheet, hdr_row)
        
## -----------------------------------------------------------------------
## -----------------------------------------------------------------------
def gen_spreadsheet(data, job_data):
    """ . """

    workbook = Workbook()
    # sheet = workbook.active

    # Where is this stray tab coming from ?
    junk = workbook.get_sheet_by_name('Sheet')
    workbook.remove_sheet(junk)
    
    workbook.create_sheet('summary')
    summary_page(workbook)

    view_map = jobs_by.IndexUtils().view_to_jobs()
    
    for rec in get_view_tabs():
        idx, vx, view = rec

        ## Cannot use literal view name, 4 tabs fill screen
        ws = workbook.create_sheet(title=vx)
        do_header(ws)

        if view not in view_map:
            print(" ** warning: view not mapped %s" % view)
            continue

        fills = get_fills()
        
        row = 5
        for job in view_map[view]:

            print(" ** view: %s, job: %s" % (view, job))

            if job not in job_data:
                print(" ** warning: JOB NOT FOUND [%s]" % job)
                continue

            # apply_to_row()
            for rec in job_data[job]:

                if False:
                    hdr        = get_header_rec('DATE')
                    cell       = ws.cell(row=row, column=hdr['col'])
                    cell.value = rec['ts_ccyymmdd']

                if True: 
                    hdr        = get_header_rec('TIME')
                    cell       = ws.cell(row=row, column=hdr['col'])
                    cell.value = rec['ts_hhmmss']

                if True:
                    hdr        = get_header_rec('DAY')
                    cell       = ws.cell(row=row, column=hdr['col'])
                    cell.value = rec['weekday']

                if rec['state'] == 'DISABLED':
                    hdr        = get_header_rec('ACTIVE')
                    col        = hdr['col']
                    cell       = ws.cell(row=row, column=col)
                    cell.value = 'N' 

                    if True:
                        cell      = ws.cell(row=row, column=col)
                        cell.fill = get_fills('grey')
                        cell.alignment = Alignment(horizontal='center',vertical='center')

                    elif False:
                        max_col    = ws.max_column
                        if 15 > max_col:
                            max_col = 15
                        for col_fill in range(1, max_col):
                            cell      = ws.cell(row=row, column=col_fill)
                            cell.fill = get_fills('grey')


                hdr  = get_header_rec('SFUA') # col=9
                col  = hdr['col']
                row  = row + 1  
                cell = ws.cell(row=row, column=col)
                cell.value = rec['result']
                if cell.value == 'ABORTED':
                    cell.fill = fills['orange']
                elif cell.value == 'FAILURE':
                    cell.fill = fills['pink']

                col = col + 1
                cell = ws.cell(row=row, column=col)
                cell.style = "Hyperlink"
                cell.value     = rec['job_id']
                cell.hyperlink = rec['urls'][view]

                if False:
                    hdr        = get_header_rec('DURATION')
                    cell       = ws.cell(row=row, column=hdr['col'])
                    cell.value = rec['duration']
                
                hdr        = get_header_rec('NOTES')
                cell       = ws.cell(row=row, column=hdr['col'])
                cell.value = rec['job_name']

    apply_per_sheet(workbook)

    return workbook

## [SEE ALSO]
## ---------------------------------------------------------------------------
# https://stackoverflow.com/questions/39077661/adding-hyperlinks-in-some-cells-openpyxl    
# https://www.soudegesu.com/en/post/python/sheet-excel-with-openpyxl/

# ws2 = hospital_ranking.create_sheet(title = 'California')
# ws2 = hospital_ranking.get_sheet_by_name('California')
## ---------------------------------------------------------------------------

# https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/hyperlink.html

# EOF
